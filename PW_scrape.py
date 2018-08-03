import urllib.request
import os
import io
import re
import sys
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup


def updt(total, progress):
    """
    Displays or updates a console progress bar.

    Original source: https://stackoverflow.com/a/15860757/1391441
    """
    barLength, status = 20, ""
    progress = float(progress) / float(total)
    if progress >= 1.:
        progress, status = 1, "\r\n"
    block = int(round(barLength * progress))
    text = "\r[{}] {:.0f}% {}".format(
        "#" * block + "-" * (barLength - block), round(progress * 100, 0),
        status)
    sys.stdout.write(text)
    sys.stdout.flush()


def create_file(filename):
    cwd = os.getcwd()
    final_directory = os.path.join(cwd, filename)
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)


def open_save_link(link, filename):
    filename = filename.rstrip()
    filename = re.sub('[^A-Za-z0-9 ]+', ' ', filename)
    filename = filename.rstrip()
    page = urllib.request.urlopen(link)
    page = BeautifulSoup(page, "lxml")
    page = page.prettify()
    create_file(filename)
    directory = os.getcwd()
    save_file = os.path.join(directory, filename)
    save_file = os.path.join(save_file, filename + ".html")
    with io.open(save_file, "w+", encoding="utf-8") as f:
        f.write(page)
        f.close()
    return page


def strip_data(bs4_object):
    arr = []
    common_name = []
    species_name = []
    symptoms = []

    soup = BeautifulSoup(bs4_object, "html.parser")
    try:
        main_content = soup.find("div", {"id": "cphMainContent_cphPWContentBody_pnlHostPlants"})
        main_content = main_content.find_all("div", {"class": "SectionItem"})
        for i in range(0, len(main_content)):
            raw_text = main_content[i].text
            strip = re.split("\((.*?)\)", raw_text)
            if len(strip) is 3:
                # print(str(i) + ": " + strip[0].lstrip() + ", " + strip[1].rstrip())
                common_name.append(strip[0].strip())
                species_name.append(strip[1].strip())
            else:
                # print(str(i) + ": " + strip[0].strip())
                common_name.append(strip[0].strip())
                species_name.append("N/A")
    except:
        common_name.append("N/A")
        species_name.append("N/A")
    arr.append(common_name)
    arr.append(species_name)

    secondary_content = soup.find("div", {"id": "cphMainContent_cphPWContentBody_pnlSymptomsList"})
    if secondary_content:
        secondary_content = secondary_content.find_all("div", {"class": "SectionItem"})
        for i in range(0, len(secondary_content)):
            raw_text = secondary_content[i]
            raw_text = raw_text.find_all("span")
            symptoms.append(raw_text[0].text.strip() + raw_text[1].text.strip() + raw_text[2].text.strip())
    arr.append(symptoms)
    return arr


def add_data(line, main_sheet, clean_sheet, data):
    start_row = clean_sheet.max_row + 1
    plant_common_name = data[0]
    plant_species_name = data[1]
    if len(data[2]) == 0:
        plant_symptoms = "N/A"
    else:
        plant_symptoms = ', '.join(data[2])
    insect_common_name = main_sheet.cell(row=line, column=1).value
    insect_scientific_name = main_sheet.cell(row=line, column=2).value

    if len(plant_common_name) != len(plant_species_name):
        print("Error: Common names not equal to Species names")
    else:
        for i in range(0, len(plant_common_name)):
            clean_sheet.cell(row=start_row, column=2).value = insect_scientific_name
            clean_sheet.cell(row=start_row, column=3).value = insect_common_name
            clean_sheet.cell(row=start_row, column=6).value = plant_species_name[i]
            clean_sheet.cell(row=start_row, column=9).value = plant_common_name[i]
            clean_sheet.cell(row=start_row, column=14).value = plant_symptoms
            start_row += 1


# load the workbook
wb = load_workbook(filename="USA_List.xlsx", data_only=True)

# Select sheets to work with
rawData = wb['Sheet1']
cleanData = wb.create_sheet(title="list")

# List of links
links = []

# Extract the links
for cell in range(2, rawData.max_row+1):
    cellLink = rawData.cell(row=cell, column=3).hyperlink.display
    links.append(cellLink)

# Open links and start processing data
for cell in range(2, len(links)+1):
    linkNumber = cell - 2
    # site_data is the BS4 object. Calling this function to download the pages, save them and then return the html data
    site_data = open_save_link(links[linkNumber], rawData.cell(row=cell, column=1).value)
    # strip_data basically strips all the useful data we're looking for and returns an array of arrays containing  data
    data_arr = strip_data(site_data)
    # add_data adds the striped data to the excel file
    add_data(cell, rawData, cleanData, data_arr)
    # this is progress bar
    time.sleep(.1)
    updt(len(links)+1, cell)

wb.save("USA_List.xlsx")



